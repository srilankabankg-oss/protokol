import os
import tempfile
from datetime import datetime, UTC
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from src.adapters.db.models import Meeting, Task
from src.core.enums import WorkflowStage


def _task_stage_label(stage: WorkflowStage) -> str:
    labels = {
        WorkflowStage.TO_DO: "К выполнению",
        WorkflowStage.IN_PROGRESS: "В работе",
        WorkflowStage.ESCALATED: "Эскалировано",
        WorkflowStage.COMPLETED: "Завершено",
    }
    return labels.get(stage, stage.value)


def generate_xlsx(meeting: Meeting, tasks: list[Task], output_dir: str = "/tmp") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Протокол"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:F1")
    ws["A1"] = f"ПРОТОКОЛ: {meeting.title}"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Дата: {meeting.date.strftime('%d.%m.%Y %H:%M') if meeting.date else '—'}"
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = f"Уровень: {meeting.level.value} | Статус: {meeting.status.value}"
    ws["A3"].alignment = Alignment(horizontal="center")

    headers = ["№", "Номер задачи", "Описание", "Статус", "Срок", "RACI"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, task in enumerate(tasks, 6):
        raci_parts = []
        for ra in task.raci_assignments:
            raci_parts.append(f"{ra.role.value}: {str(ra.user_id)[:8]}")

        row_data = [
            row_idx - 5,
            task.task_number,
            task.description[:200],
            _task_stage_label(task.workflow_stage),
            task.deadline.strftime("%d.%m.%Y") if task.deadline else "—",
            "; ".join(raci_parts) if raci_parts else "—",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 30

    filename = f"protocol_{meeting.id}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


def generate_pdf(meeting: Meeting, tasks: list[Task], output_dir: str = "/tmp") -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.lib import colors
    from reportlab.platypus.flowables import HRFlowable
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    filename = f"protocol_{meeting.id}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(output_dir, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("Title_RU", parent=styles["Title"], fontSize=16, spaceAfter=6*mm)
    heading_style = ParagraphStyle("Heading_RU", parent=styles["Heading2"], fontSize=13, spaceBefore=8*mm, spaceAfter=3*mm)
    body_style = ParagraphStyle("Body_RU", parent=styles["Normal"], fontSize=10, leading=14)

    story = []

    story.append(Paragraph(f"ПРОТОКОЛ СОВЕЩАНИЯ", title_style))
    story.append(Paragraph(f"<b>{meeting.title}</b>", styles["Heading2"]))
    story.append(Paragraph(
        f"Дата: {meeting.date.strftime('%d.%m.%Y %H:%M') if meeting.date else '—'} | "
        f"Уровень: {meeting.level.value} | Статус: Утверждено",
        body_style,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2563EB")))
    story.append(Spacer(1, 5*mm))

    if meeting.content_markdown:
        story.append(Paragraph("СОДЕРЖАНИЕ ПРОТОКОЛА", heading_style))
        for line in meeting.content_markdown.split("\n"):
            if line.startswith("## "):
                story.append(Paragraph(f"<b>{line[3:]}</b>", heading_style))
            elif line.strip():
                story.append(Paragraph(line, body_style))

    if tasks:
        story.append(Spacer(1, 5*mm))
        story.append(Paragraph("ПОРУЧЕНИЯ", heading_style))

        table_data = [["№", "Номер", "Описание", "Статус"]]
        for i, task in enumerate(tasks, 1):
            table_data.append([
                str(i),
                task.task_number,
                task.description[:100],
                _task_stage_label(task.workflow_stage),
            ])

        col_widths = [20*mm, 35*mm, 90*mm, 30*mm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)

    story.append(Spacer(1, 10*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(Paragraph(
        f"Документ сгенерирован автоматически платформой «Книга добрых дел» "
        f"{datetime.now(UTC).strftime('%d.%m.%Y %H:%M')} UTC",
        ParagraphStyle("Footer", parent=body_style, fontSize=7, textColor=colors.grey),
    ))

    doc.build(story)
    return filepath