import io


async def parse_uploaded_file(content: bytes, ext: str) -> str:
    if ext == "txt" or ext == "md":
        return content.decode("utf-8", errors="replace")

    if ext == "docx":
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            return content.decode("utf-8", errors="replace")

    if ext == "xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"=== {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join(str(c) if c is not None else "" for c in row)
                    if row_text.strip():
                        lines.append(row_text)
            return "\n".join(lines)
        except ImportError:
            return content.decode("utf-8", errors="replace")

    raise ValueError(f"Unsupported format: {ext}")